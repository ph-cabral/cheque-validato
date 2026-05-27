import os
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def validar_cuit(cuit: str) -> bool:
    if len(cuit) != 11 or not cuit.isdigit():
        return False
    pesos = [5,4,3,2,7,6,5,4,3,2]
    s = sum(int(cuit[i]) * pesos[i] for i in range(10))
    return ((11 - (s % 11)) % 11) == int(cuit[-1])

def generar_cheques_rechazados(cuits):
    wb = Workbook()
    ws = wb.active
    ws.append(["CUIT", "NOMBRE", "NUM CHEQUE", "FECHA RECHAZO", "MONTO"])

    base_url = "https://api.bcra.gob.ar/CentralDeDeudores/v1.0/Deudas/ChequesRechazados"
    colors = ["FFFFDF20", "FF9AE630", "FF7C86FF"]
    color_index = 0

    # Session con headers para simular navegador
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
        "Accept-Language": "es-AR,es;q=0.9",
        "Connection": "keep-alive"
    })
    session.verify = False

    for cuit in cuits:
        cuit = str(int(cuit)).strip()
        intentos = 3  # reintentos por si falla
        print(cuit)
        while intentos > 0:
            try:
                r = session.get(
                    f"{base_url}/{cuit}",
                    timeout=5
                )

                if r.status_code == 404:
                    print(f"CUIT {cuit}: sin datos")
                    break
                elif r.status_code != 200:
                    print(f"CUIT {cuit}: status {r.status_code}")
                    break

                data = r.json().get("results", {})
                fill = PatternFill(
                    start_color=colors[color_index % 3],
                    end_color=colors[color_index % 3],
                    fill_type="solid"
                )

                for causal in data.get("causales", []):
                    for entidad in causal.get("entidades", []):
                        for det in entidad.get("detalle", []):
                            ws.append([
                                data.get("identificacion", ""),
                                data.get("denominacion", ""),
                                det.get("nroCheque", ""),
                                det.get("fechaRechazo", ""),
                                det.get("monto", "")
                            ])
                            for cell in ws[ws.max_row]:
                                cell.fill = fill

                color_index += 1
                print(r.status_code)
                break  # salir del while si fue exitoso

            except requests.exceptions.ConnectionError as e:
                intentos -= 1
                print(f"Conexión cortada para CUIT {cuit}. Reintentos restantes: {intentos}")
                if intentos > 0:
                    time.sleep(5)  # esperar más antes de reintentar
                else:
                    print(f"CUIT {cuit} omitido por error de conexión")
                print("nope1")

            except requests.exceptions.Timeout:
                print(f"Timeout para CUIT {cuit}, omitiendo...")
                print("nope2")
                break

            except Exception as e:
                print(f"Error inesperado para CUIT {cuit}: {e}")
                print("nope3")
                break

        time.sleep(3)  # pausa entre cada CUIT

    wb.save("../cheques_Rechazados.xlsx")
    print("Archivo guardado: cheques_Rechazados.xlsx")
