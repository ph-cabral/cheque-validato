import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# def limpiar_fecha(valor):
#     if pd.isna(valor):
#         return pd.NaT
#     if isinstance(valor, (int, float)):
#         try:
#             return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(valor), unit='D')
#         except:
#             return pd.NaT
#     if isinstance(valor, str):
#         for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
#             try:
#                 return pd.to_datetime(datetime.strptime(valor.strip(), fmt))
#             except:
#                 pass
#         return pd.to_datetime(valor, dayfirst=True, errors='coerce')
#     if isinstance(valor, datetime):
#         return pd.to_datetime(valor)
#     return pd.NaT

def limpiar_fecha(valor):
    """
    Convierte cualquier formato de fecha a datetime de pandas.
    Maneja: int/float (Excel serial), strings, datetime objects.
    """
    # Caso 1: Valor nulo
    if pd.isna(valor):
        return pd.NaT
    
    # Caso 2: Número (fecha serial de Excel)
    if isinstance(valor, (int, float)):
        try:
            return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(valor), unit='D')
        except:
            return pd.NaT
    
    # Caso 3: Ya es datetime
    if isinstance(valor, datetime):
        return pd.to_datetime(valor)
    
    # Caso 4: String - probar formatos específicos primero
    if isinstance(valor, str):
        valor = valor.strip()
        
        # Formatos explícitos (más rápido y sin warnings)
        formatos = [
            "%d/%m/%Y",           # 25/12/2024
            "%d-%m-%Y",           # 25-12-2024
            "%Y-%m-%d",           # 2024-12-25
            "%Y-%m-%d %H:%M:%S",  # 2024-12-25 14:30:00
            "%Y-%m-%d %H:%M:%S.%f",  # 2024-12-25 14:30:00.123456
            "%d/%m/%Y %H:%M:%S",  # 25/12/2024 14:30:00
        ]
        
        for fmt in formatos:
            try:
                return pd.to_datetime(datetime.strptime(valor, fmt))
            except (ValueError, TypeError):
                continue
        
        # Si ninguno funciona, dejar que pandas infiera (SIN dayfirst)
        try:
            return pd.to_datetime(valor, errors='coerce')
        except:
            return pd.NaT
    
    # Caso por defecto
    return pd.NaT


def ajustar_ancho_excel(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(nombre_archivo)


from openpyxl.styles import PatternFill, Font

def formatear_excel(nombre_archivo, col_importe='Importe'):
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    idx_importe = None
    for cell in ws[1]:
        if cell.value == col_importe:
            idx_importe = cell.column
            break

    fill_subtotal = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font_subtotal = Font(bold=True)

    if idx_importe:
        letra = get_column_letter(idx_importe)
        for cell in ws[letra][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

        # Colorear filas de subtotal: solo tienen valor en Importe
        for row in ws.iter_rows(min_row=2):
            otras = [c.value for c in row if c.column != idx_importe]
            importe_val = row[idx_importe - 1].value
            if all(v is None for v in otras) and isinstance(importe_val, (int, float)):
                for cell in row:
                    cell.fill = fill_subtotal
                    cell.font = font_subtotal

    for col in ws.columns:
        max_length = 0
        letra = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            if cell.column == idx_importe and isinstance(cell.value, (int, float)):
                texto = f"${cell.value:,.2f}"
            else:
                texto = str(cell.value)
            max_length = max(max_length, len(texto))
        ws.column_dimensions[letra].width = max_length + 2

    wb.save(nombre_archivo)